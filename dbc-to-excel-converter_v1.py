from tkinter import *
from tkinter import filedialog, messagebox
import re
import os

from openpyxl import Workbook
from openpyxl.worksheet.table import Table
import cantools

inp= ""
out= ""

def setTextInput(text):
    e1.delete(0,"end")
    e1.insert(0, text)
    op_name = re.split("/", text)[-1].replace('.dbc', '.xlsx')
    e2.insert(0, op_name)
    return op_name

def resetTextInput():
    e1.delete(0,"end")
    e2.delete(0,"end")

def browseFiles():
    try:
        filename = filedialog.askopenfilename(initialdir="/", title="Select a DBC File",
                                              filetypes=(("DBC files", "*.dbc*"), ("all files", "*.*")))
        op_name = setTextInput(filename)

        global inp, out
        inp = filename
        out = op_name

    except:
        messagebox.showerror("Error", "Unable to load File")


def convert_to_excel():
    try:
        global inp, out
        out2=out
        temp_str = out2
        op_name_from_e2 = e2.get()
        if op_name_from_e2 != temp_str:
            out=out2.replace(temp_str, op_name_from_e2)

        db = cantools.database.load_file(inp)

        wb = Workbook()
        ws = wb.create_sheet('messages')
        ws.auto_filter.ref = ws.dimensions

        sheet = wb["Sheet"]
        wb.remove(sheet)

        data1 = []
        for msg in db.messages:
            data1.append(['0x'+(format(int(msg.frame_id), 'X')), msg.name, msg.comment, msg.length, msg.cycle_time, msg.send_type,
                          str(int(msg.cycle_time) * 10) if int(msg.cycle_time) < 100 else str(int(msg.cycle_time) * 5),
                          ' ', ' ',' ', ' ',msg.frame_id,
                          re.sub('\[|\]|\'', '', str(msg.senders)), ' ', ' ', ' ', ' ', ' '])
        ws.append(['Message ID [hex]', 'Message Name', 'Message Description', 'Data Length Code',
                   'Cycle Time [ms]','Send Mode', 'Timeout [ms]', 'Delay Time [ms]',
                   'Msg Start Delay Time [ms]', 'Cycle Time Fast [ms]','Number of Repetition',
                   'Message ID [dez]','Transmitter','Diag Request','Diag Response','DiagState',
                   'AutoSAR NM', 'IL Support'])

        ws.freeze_panes = "A2"
        ws.print_title_rows = '1:1'

        for row in data1:
            ws.append(row)
        ref_str = "A1" + ":" + "I" + str(len(data1) + 1)
        tab1 = Table(displayName="Table1", ref=ref_str)

        ws.add_table(tab1)

        ws = wb.create_sheet('signals')
        ws.auto_filter.ref = ws.dimensions

        data2 = []
        for msg in db.messages:
            for sig in msg.signals:
                val_str = ""
                if (str(sig.choices) != 'None'):
                    for k, v in sig.choices.items():
                        val_str += str(k) + ' \"' + str(v) + '\" '
                data2.append(
                    [str(msg.bus_name)+'-CAN', sig.name, sig.comment,msg.name, '0x'+(format(int(msg.frame_id), 'X')), sig.length, sig.start,
                     "Motorola" if sig.byte_order == "big_endian" else "Intel",
                     "Signed" if sig.is_signed == "FALSE" else "Unsigned",
                     sig.scale, sig.offset,
                     sig.minimum,
                     sig.maximum,sig.initial, sig.unit, msg.cycle_time, msg.send_type, msg.send_type,
                     val_str if str(sig.choices)!='None' else 'n.a',
                     ' ',
                     str(int(msg.cycle_time)*10) if int(msg.cycle_time)<100 else str(int(msg.cycle_time)*5),
                     '1' if str(sig.is_multiplexer)=='FALSE' else '0',
                     re.sub('\[|\]|\'', '', str(msg.senders)),
                     re.sub('\[|\]|\'','',  str(sig.receivers)),
                     '-',' '])
        ws.append(['Bus', 'Signal Name', 'Signal Description', 'Message Name', 'Message ID [Dec]', 'Signal length', 'Start Bit',  'Byte Order', 'Sign',  'Factor', 'Offset', 'minimum',
                   'maximum','initial', 'unit','Cycle Time [ms]','Signal Send Mode','Message Send Mode', 'Value Matrix', 'Invalid Value [Hex]','Timeout Signal [ms]','Multiplex Value [dez]', 'Senders','Receivers', 'Timeout Value [Hex]','Comments'])

        ws.freeze_panes = "A2"
        ws.print_title_rows = '1:1'

        for row in data2:
            ws.append(row)
        ref_str = "A1" + ":" + "I" + str(len(data2) + 1)
        tab2 = Table(displayName="Table2", ref=ref_str)

        ws.add_table(tab2)
        wb.save(os.path.expanduser('~\\Documents\\'+out))
        resetTextInput()
        messagebox.showinfo("Export Successful", "File saved to "+os.path.expanduser('~\\Documents\\'+out))

        print("Export successful")

    except:
        print("failed")
        resetTextInput()
        messagebox.showerror("Error", "Unable to Export!!")

master = Tk()
master.title("DBC to Excel Converter")
master.minsize(width=400, height=200)

l1=Label(master, text="Input File Path").grid(row=0)
l2=Label(master, text="Output File Name").grid(row=1)

e1 = Entry(master, width=30)
e2 = Entry(master, width=30)

button_explore = Button(master, text="Browse File", command=browseFiles).grid(row=0, column=2)

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)

Button(master, text='Convert',command=convert_to_excel).grid(row=3, column=1,pady=4)
Button(master, text='Reset', command=resetTextInput).grid(row=3, column=1, sticky=W,pady=4)

mainloop()
