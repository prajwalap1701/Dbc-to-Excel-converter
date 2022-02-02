# CAN DBC to Excel converter v3.6

from tkinter import *
from tkinter import filedialog, messagebox
import re
import os
import csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
import cantools

output_file_name= ""
msg_csv_file_name = ''
sig_csv_file_name = ''
wb = Workbook()
db = ''

def setTextInput(text):
    e1.delete(0,"end")
    e1.insert(0, text.replace('/', '\\'))
    op_name = re.split("/", text)[-1].replace('.dbc', '.xlsx')
    e4.insert(0, op_name)
    return op_name

def resetTextInput():
    e1.delete(0, "end")
    e2.delete(0, "end")
    e3.delete(0, "end")
    e4.delete(0, "end")

def browse_msg_csv():
    try:
        e2.delete(0, "end")
        filepath = filedialog.askopenfilename(initialdir="/", title="Select a Message CSV File",
                                              filetypes=(("CSV files", "*.csv*"), ("all files", "*.*")))
        global msg_csv_file_name
        msg_csv_file_name = filepath.replace('/', '\\')
        e2.insert(0, msg_csv_file_name)
    except:
        messagebox.showerror("Error", "Unable to browse Message CSV")
        print('error')

def browse_sig_csv():
    try:
        e3.delete(0, "end")
        filepath = filedialog.askopenfilename(initialdir="/", title="Select a Signal CSV File",
                                              filetypes=(("CSV files", "*.csv*"), ("all files", "*.*")))
        global sig_csv_file_name
        sig_csv_file_name = filepath.replace('/', '\\')
        e3.insert(0, sig_csv_file_name)
    except:
        messagebox.showerror("Error", "Unable to browse Signal CSV")
        print('error')

def browseFiles():
    try:
        filename = filedialog.askopenfilename(initialdir="/", title="Select a DBC File",
                                              filetypes=(("DBC files", "*.dbc*"), ("all files", "*.*")))
        op_name = setTextInput(filename)
        global output_file_name, db
        db = cantools.database.load_file(filename)
        output_file_name = op_name

    except:
        messagebox.showerror("Error", "Unable to load File")


def write_excel():
    try:
        global output_file_name, msg_csv_file_name, sig_csv_file_name
        out2=output_file_name
        temp_str = out2
        op_name_from_e2 = e4.get()
        if op_name_from_e2 != temp_str:
            output_file_name=out2.replace(temp_str, op_name_from_e2)

        if msg_csv_file_name == '':
            msg_csv_file_name = e2.get()

        if sig_csv_file_name == '':
            sig_csv_file_name = e3.get()

        with open(msg_csv_file_name, newline='') as f:
            reader = csv.reader(f)
            msg_data = list(reader)

        with open(sig_csv_file_name, newline='') as f:
            reader = csv.reader(f)
            sig_data = list(reader)

        if 'Comment' in sig_data[0]:
            messagebox.showerror("Error", "Comment column found in Signal CSV file!")
            exit(0)

        try:
            diagrequest_index = msg_data[0].index('DiagRequest')
            diagresponse_index = msg_data[0].index('DiagResponse')
            diagstate_index = msg_data[0].index('DiagState')
            delaytime_index = msg_data[0].index('GenMsgDelayTime')
            startdelaytime_index = msg_data[0].index('GenMsgStartDelayTime')
            cycletimefast_index = msg_data[0].index('GenMsgCycleTimeFast')
            numofrepition_index = msg_data[0].index('GenMsgNrOfRepetition')
            nmasr_index = msg_data[0].index('NmAsrMessage')
            ilsupport_index = msg_data[0].index('GenMsgILSupport')
            print("Message csv file format is correct")

        except:
            print("Message csv file format is not correct")
            messagebox.showerror("Error", "Message CSV file Invalid!")

        try:
            invalidval_index = sig_data[0].index('InvalidValue')
            sigsendtype_index = sig_data[0].index('GenSigSendType')
            if 'GenSigTimeoutTime' in sig_data[0] and 'GenSigTimeoutTime_GW' in sig_data[0]:
                sigtimeout_index = sig_data[0].index('GenSigTimeoutTime')
            else:
                sigtimeout_index = sig_data[0].index('GenSigTimeoutTime_ALL')
            max_index = sig_data[0].index('Maximum')
            min_index = sig_data[0].index('Minimum')
            initvalue_index = sig_data[0].index('Initial Value')
            unit_index = sig_data[0].index('Unit')
            print("Signals csv file format is correct")

        except:
            print("Signals csv file format is not correct")
            messagebox.showerror("Error", "Signals CSV file Invalid!")

        msg_list_data = msg_data[1:]
        sig_list_data = sig_data[1:]

        # global wb

        # for updating several fields in messages tab
        messages_sheet = wb["messages"]

        row_count_msg = 1
        for row in messages_sheet.iter_cols(min_row=2, min_col=8, max_col=8):
            for cell in row:
                for list_row in msg_list_data:
                    if cell.value == list_row[0]:
                        row_count_msg += 1
                        messages_sheet['N' + str(row_count_msg)] = int(list_row[delaytime_index])
                        messages_sheet['O' + str(row_count_msg)] = int(list_row[startdelaytime_index])
                        messages_sheet['P' + str(row_count_msg)] = int(list_row[cycletimefast_index])
                        messages_sheet['Q' + str(row_count_msg)] = int(list_row[numofrepition_index])
                        messages_sheet['T' + str(row_count_msg)] = list_row[diagrequest_index]
                        messages_sheet['U' + str(row_count_msg)] = list_row[diagresponse_index]
                        messages_sheet['V' + str(row_count_msg)] = list_row[diagstate_index]
                        messages_sheet['W' + str(row_count_msg)] = list_row[nmasr_index]
                        messages_sheet['X' + str(row_count_msg)] = list_row[ilsupport_index]

                        messages_sheet['D' + str(row_count_msg)] = '=ROUND(IF($A${0}=0,0.000,(($J${0}*8+47)*1.05*100*1000)/(500000*$A${0})),3)'.format(row_count_msg)
                        messages_sheet['E' + str(row_count_msg)] = '=ROUND(IF($A${0}=0,0.000,(($J${0}*8+47)*1.05*100*1000)/(500000*$A${0})),3)'.format(row_count_msg)
                        messages_sheet['F' + str(row_count_msg)] = '=ROUND(IF($A${0}=0,0.000,(($J${0}*8+47)*1.05*100*1000)/(500000*$A${0})),3)'.format(row_count_msg)
                        break

        # for updating SignalSendType, Timeout and Invalid value

        signals_sheet = wb["signals"]
        row_count_sig = 1
        for row in signals_sheet.iter_cols(min_row=2, min_col=2, max_col=2):
            for cell in row:
                for list_row in sig_list_data:
                    if cell.value == list_row[0]:
                        row_count_sig += 1
                        signals_sheet['L' + str(row_count_sig)] = float(list_row[min_index])
                        signals_sheet['M' + str(row_count_sig)] = float(list_row[max_index])
                        signals_sheet['N' + str(row_count_sig)] = float(list_row[initvalue_index])

                        if len(str(list_row[unit_index]).strip()) == 0:
                            signals_sheet['O' + str(row_count_sig)] = "-"
                        else:
                            signals_sheet['O' + str(row_count_sig)] = list_row[unit_index]

                        signals_sheet['Q' + str(row_count_sig)] = list_row[sigsendtype_index]

                        if len(str(list_row[invalidval_index]).strip()) == 0:
                            signals_sheet['T' + str(row_count_sig)] = "-"
                        else:
                            signals_sheet['T' + str(row_count_sig)] = list_row[invalidval_index]
                        signals_sheet['U' + str(row_count_sig)] = int(list_row[sigtimeout_index])
                        break

        # for updating timeout in message tab

        values = []
        msg_id_list = []

        for i in range(2, signals_sheet.max_row + 1):
            if signals_sheet.cell(row=i, column=5).value in msg_id_list:
                pass  # if already in list do nothing
            else:
                msg_id_list.append(signals_sheet.cell(row=i, column=5).value)
                values.append([signals_sheet.cell(row=i, column=5).value, signals_sheet.cell(row=i, column=21).value])

        messages_sheet = wb["messages"]
        row_count_tout = 1
        fid_present = False
        for row in messages_sheet.iter_cols(min_row=2, min_col=7, max_col=7):
            for cell in row:
                    for val in values:
                        if cell.value == val[0]:
                            fid_present = True
                            row_count_tout += 1
                            messages_sheet['M' + str(row_count_tout)] = int(val[1])
                    if fid_present == False:
                        row_count_tout += 1
                        messages_sheet['M' + str(row_count_tout)] = "Empty Frame"
                    fid_present = False

        wb.save(os.path.expanduser('~\\Documents\\' + output_file_name))
        resetTextInput()
        messagebox.showinfo("Export Successful",
                            "File saved to " + os.path.expanduser('~\\Documents\\' + output_file_name))
        print('stage 1.5 successful')
        return

    except:
        messagebox.showerror("Error", "Unable to Fill Excel!! Please check for consistency errors in dbc file")
        print('failed at stage 2')
        resetTextInput()


def convert_to_excel():
    try:

        global db, wb

        ws = wb.create_sheet('messages')
        # ws.auto_filter.ref = ws.dimensions

        sheet = wb["Sheet"]
        wb.remove(sheet)
        data1 = []
        for msg in db.messages:
            data1.append([ msg.cycle_time, msg.cycle_time, msg.cycle_time, '', '', '',
                           '0x'+(format(int(msg.frame_id), 'X')),
                           msg.name, str(msg.comment).strip() if msg.comment != None else '' ,
                           msg.length, msg.cycle_time, msg.send_type,
                          ' ', ' ', ' ',' ', ' ',msg.frame_id,
                          re.sub('\[|\]|\'', '', str(msg.senders)), ' ', ' ', ' ', ' ', ' '])
        ws.append(['Cycletime max. [ms]','Cycletime typ. [ms]','Cycletime min. [ms]','Busload max. %','Busload typ. %', 'Busload min. %', 'Message ID [hex]', 'Message Name', 'Message Description', 'Data Length Code',
                   'Cycle Time [ms]','Send Mode', 'Timeout [ms]', 'Delay Time [ms]',
                   'Msg Start Delay Time [ms]', 'Cycle Time Fast [ms]','Number of Repetition',
                   'Message ID [dec]','Transmitter','Diag Request','Diag Response','DiagState',
                   'AutoSAR NM', 'IL Support'])


        for row in data1:
            ws.append(row)
        ref_str = "A1" + ":" + "X" + str(len(data1) + 1)
        # tab1 = Table(displayName="Table1", ref=ref_str)
        #
        # ws.add_table(tab1)
        ws.auto_filter.ref = ws.dimensions

        ws = wb.create_sheet('signals')

        data2 = []
        for msg in db.messages:
            for sig in msg.signals:
                val_str = ""
                if (str(sig.choices) != 'None'):
                    for k, v in sig.choices.items():
                        val_str += str(k) + ' \"' + str(v) + '\" '
                data2.append(
                    [str(msg.bus_name)+'-CAN', sig.name, sig.comment,msg.name,
                     '0x'+(format(int(msg.frame_id), 'X')),
                     sig.length, sig.start,
                     "Motorola" if sig.byte_order == "big_endian" else "Intel",
                     "Signed" if sig.is_signed == "FALSE" else "Unsigned",
                     sig.scale, sig.offset,
                     sig.minimum if len(str(sig.minimum)) > 0 else '0' ,
                     sig.maximum if len(str(sig.maximum)) > 0 else '0' ,
                     sig.initial if len(str(sig.initial)) > 0 else '0',
                     sig.unit if len(str(sig.unit).strip()) > 0 else '-',
                     msg.cycle_time,
                     ' ', msg.send_type,
                     val_str.strip() if str(sig.choices)!='None' else 'n.a',
                     ' ', ' ',
                     1 if sig.is_multiplexer else 0,
                     re.sub('\[|\]|\'', '', str(msg.senders)),
                     re.sub('\[|\]|\'','',  str(sig.receivers)),
                     '-',''])
        ws.append(['Bus', 'Signal Name', 'Signal Description', 'Message Name', 'Message ID [hex]', 'Signal length', 'Start Bit',  'Byte Order', 'Sign',  'Factor', 'Offset', 'minimum',
                   'maximum','initial', 'unit','Cycle Time [ms]','Signal Send Mode','Message Send Mode', 'Value Matrix', 'Invalid Value [Hex]','Timeout Signal [ms]','Multiplex Value [dec]', 'Senders','Receivers', 'Timeout Value [Hex]','Comments'])


        for row in data2:
            ws.append(row)
        ref_str = "A1" + ":" + "Z" + str(len(data2) + 1)
        # tab2 = Table(displayName="Table2", ref=ref_str)
        #
        # ws.add_table(tab2)
        ws.auto_filter.ref = ws.dimensions

        print("Stage 1 successful")

        write_excel()


        print("Stage 2 successful")
        os._exit(0)

    except:
        print("failed at stage 1")
        resetTextInput()
        messagebox.showerror("Error", "Unable to Export!! Please check for consistency errors in dbc file")

master = Tk()
master.title("DBC to Excel Converter")
master.minsize(width=400, height=200)

l1=Label(master, text="Input CAN DBC File").grid(row=0)
l2=Label(master, text="Message CSV File").grid(row=1)
l3=Label(master, text="Signals CSV File").grid(row=2)
l4=Label(master, text="Output File Name").grid(row=3)

e1 = Entry(master, width=50)
e2 = Entry(master, width=50)
e3 = Entry(master, width=50)
e4 = Entry(master, width=50)

button_explore = Button(master, text="Browse File", command=browseFiles).grid(row=0, column=2)
browse_msg_csv = Button(master, text="Browse File", command= browse_msg_csv).grid(row=1, column=2)
browse_sig_csv = Button(master, text="Browse File", command= browse_sig_csv).grid(row=2, column=2)

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)
e3.grid(row=2, column=1)
e4.grid(row=3, column=1)

Button(master, text='Convert',command=convert_to_excel).grid(row=4, column=1,pady=4)
Button(master, text='Reset', command=resetTextInput).grid(row=4, column=1, sticky=W,pady=4)

mainloop()